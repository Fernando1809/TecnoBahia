from pathlib import Path
import re, base64
from io import BytesIO
import openpyxl

p = Path('previwe.html')
text = p.read_text(encoding='utf-8')
i = text.find('DEFAULT_PEDIDO_TEMPLATE_BASE64')
print('pos', i)
j = text.find('"', i)
k = text.find('"', j+1)
print('quote', j, k)
base64str = text[j+1:k]
print('len', len(base64str))
data = base64.b64decode(base64str)
print('decoded', len(data))
wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
print('sheets', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.strip().upper() in ('SUB-TOTAL =', 'IVA =', 'TOTAL ='):
                print('label', name, cell.coordinate, cell.value, ws.cell(row=cell.row, column=cell.column+1).value)
