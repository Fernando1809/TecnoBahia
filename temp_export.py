from pathlib import Path
import re, base64
from io import BytesIO
import openpyxl
import json

p = Path('previwe.html')
text = p.read_text(encoding='utf-8')
i = text.find('DEFAULT_PEDIDO_TEMPLATE_BASE64')
j = text.find('"', i)
k = text.find('"', j+1)
base64str = text[j+1:k]
data = base64.b64decode(base64str)
wb = openpyxl.load_workbook(BytesIO(data), data_only=False)

pedido_sheet_name = 'PEDIDO' if 'PEDIDO' in wb.sheetnames else wb.sheetnames[0]
ws = wb[pedido_sheet_name]

out = Path('temp_export.xlsx')
wb.save(out)
print('saved', out)
# try read back
wb2 = openpyxl.load_workbook(out, data_only=False)
print('reloaded', wb2.sheetnames)
